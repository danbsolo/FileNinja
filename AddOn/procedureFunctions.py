from openpyxl.styles import Font, PatternFill
# from openpyxl.utils import get_column_letter
import os
from FileNinjaSuite.Shared.sharedProcedureHelpers import *



def nameChopAIQueryFunction(inColIndex, outColIndex, firstRow, lastRow, inputList, outputList, ws):
    i = 0
    lastRowYielded = firstRow + len(outputList) -1
    
    ## If these don't equal, then it's the AI that likely messed up
    # if lastRowYielded != lastRow:
    #     print(f"ERROR: LastRow = {lastRow} is not equal to LastRowYielded = {lastRowYielded}. outputList is unexpected length.")
    # else:
    #     print("outputList is expected length.")
    
    # Using lastRowYielded makes this at least avoid IndexOutOfBoundsError
    for cellRow in range(firstRow, lastRowYielded+1):
        ws.cell(row=cellRow, column=outColIndex, value=outputList[i]) # col E
        i += 1



def nameChopModifierFunction(inColIndex, outColIndex, firstRow, lastRow, inputList, outputList, ws):
    """Slight bug in that a file can be successfully 'renamed' into the same text but with different cases,
    even if it wasn't actually modified."""

    # Always assumes the directory index is column 0
    dirColIndex = 1 # inColIndex -1
    # oldFilenameColIndex = inColIndex
    # newFilenameColIndex = outColIndex

    # define font stuff
    boldFont = Font(bold=True) # grayish
    successFill = PatternFill(fill_type="solid", fgColor="00FF80")
    failureFill = PatternFill(fill_type="solid", fgColor="FF4444")

    # get the first directory to be used    
    currentDirectory = ""
    i = firstRow
    while not currentDirectory:
        if (potentialDirectory := ws.cell(row=i, column=dirColIndex).value):
            currentDirectory = addLongPathPrefix(potentialDirectory)
        i -= 1


    for cellRow in range(firstRow, lastRow+1):
        # if we're on a new directory now, update currentDirectory
        if (potentialDirectory := ws.cell(row=cellRow, column=dirColIndex).value):
            currentDirectory = addLongPathPrefix(potentialDirectory)

        oldFilenameCell = ws.cell(row=cellRow, column=inColIndex)
        newFilenameCell = ws.cell(row=cellRow, column=outColIndex)

        # if oldFilenameCell and newFilenameCell are both not none and not identical, make the change
        if (oldFilenameCell.value and newFilenameCell.value) and (oldFilenameCell.value != newFilenameCell.value):
            try:
                os.rename(
                    f"{currentDirectory}\\{oldFilenameCell.value}",
                    f"{currentDirectory}\\{newFilenameCell.value}"
                )

                newFilenameCell.font = boldFont
                newFilenameCell.fill = successFill

            except Exception as e:
                newFilenameCell.font = boldFont
                newFilenameCell.fill = failureFill
        
        # write some sort of indicator the status of the rename in the new excel sheet


# TODO: Lots of repeated code between this and the TREE version. Fix that.
def fileShredFlatFunction(inColIndex, outColIndex, firstRow, lastRow, inputList, outputList, ws):
    # inColIndex is the fileNames column
    # directory is assumed to be to the right of this index
    # outColIndex is the "stage for deletion" column
    fileColIndex = inColIndex
    dirColIndex = inColIndex +1
    deletionColIndex = outColIndex

    successFill = PatternFill(fill_type="solid", fgColor="00FF80") # green
    failureFill = PatternFill(fill_type="solid", fgColor="FF4444") # red

    for cellRow in range(firstRow, lastRow+1):
        # If the value is either empty or 0, skip
        if not ws.cell(row=cellRow, column=deletionColIndex).value:
            continue
        
        fullPath = addLongPathPrefix(joinDirToFileName(
            ws.cell(row=cellRow, column=dirColIndex).value, ws.cell(row=cellRow, column=fileColIndex).value))
        
        try:
            os.remove(fullPath)
            ws.cell(row=cellRow, column=deletionColIndex).fill = successFill
        except:
            ws.cell(row=cellRow, column=deletionColIndex).fill = failureFill



def fileShredTreeFunction(inColIndex, outColIndex, firstRow, lastRow, inputList, outputList, ws):
    fileColIndex = inColIndex
    dirColIndex = 1
    deletionColIndex = outColIndex

    successFill = PatternFill(fill_type="solid", fgColor="00FF80") # green
    failureFill = PatternFill(fill_type="solid", fgColor="FF4444") # red

    # get the first directory to be used    
    currentDirectory = ""
    i = firstRow
    while not currentDirectory:
        if (potentialDirectory := ws.cell(row=i, column=dirColIndex).value):
            currentDirectory = addLongPathPrefix(potentialDirectory)
        i -= 1

    for cellRow in range(firstRow, lastRow+1):
        if (potentialDirectory := ws.cell(row=cellRow, column=dirColIndex).value):
            currentDirectory = addLongPathPrefix(potentialDirectory)

        # If the value is either empty or 0, skip
        if not ws.cell(row=cellRow, column=deletionColIndex).value:
            continue
        
        fullPath = joinDirToFileName(
            currentDirectory, ws.cell(row=cellRow, column=fileColIndex).value)
        
        try:
            os.remove(fullPath)
            ws.cell(row=cellRow, column=deletionColIndex).fill = successFill
        except Exception as e:
            ws.cell(row=cellRow, column=deletionColIndex).fill = failureFill
        


def typoFixerFunction(inColIndex, outColIndex, firstRow, lastRow, inputList, outputList, ws):
    # write
    i = 0
    for cellRow in range(firstRow, lastRow+1):
        ws.cell(row=cellRow, column=outColIndex, value=outputList.outputList[i])
        i += 1
