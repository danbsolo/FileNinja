import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import ttk
from FileNinjaSuite.FileChop import control
import openpyxl as opxl
from FileNinjaSuite.FileChop.defs import *
from FileNinjaSuite.Shared import guiController
from FileNinjaSuite.Shared import sharedCommon
import threading

def launchView():
    def launchControllerWorker():
        guiC.setStatusRunning()
        currentStatusPair = control.launchController(
            excelPathVar.get(),
            wb,
            worksheetCombobox.get(),
            procedureCombobox.get(),
            inputColEntryVar.get(),
            outputColEntryVar.get(),
            firstRowVar.get(),
            lastRowVar.get()
        )
        guiC.setCurrentStatus(currentStatusPair[0], currentStatusPair[1])
        
        
    def launchControllerFromView():
        root.title(TITLE + ": RUNNING...")
        executeButton.config(text="RUNNING....", state="disabled")

        executionThread = threading.Thread(target=launchControllerWorker)
        executionThread.daemon = True  # When the main thread closes, this daemon thread will also close alongside it
        executionThread.start()

        scheduleCheckIfDone(executionThread)

    def scheduleCheckIfDone(t):
        root.after(500, checkIfDone, t)

    def checkIfDone(t):
        # If the thread has finished
        if not t.is_alive():
            root.title(TITLE)
            executeButton.config(text="Execute", state="normal")

            exitCode, _ = guiC.getCurrentStatus()
            if exitCode == STATUS_SUCCESSFUL:
                return          
            tk.messagebox.showerror(f"Error: {exitCode}", sharedCommon.interpretError(guiC.getCurrentStatus()))
            guiC.setStatusIdle()
        else:
            scheduleCheckIfDone(t)


    def selectExcel():
        excelPath = askopenfilename(title="Select an EXCEL workbook",
                            filetypes=[("Excel files", "*.xlsx *.xlsm")])
        
        if not excelPath: return
        
        excelPathVar.set(excelPath)

        nonlocal wb
        wb = opxl.load_workbook(excelPath)

        # Update the combobox accordingly
        worksheetCombobox.config(values=wb.sheetnames)
        worksheetCombobox.set(wb.sheetnames[0])
        worksheetCombobox.event_generate("<<ComboboxSelected>>") # or just call the function directly


    def updateMaxRow():
        lastRowVar.set(wb[worksheetCombobox.get()].max_row)

    def updateColumns():
        columns = PROCEDURES_DISPLAY[procedureCombobox.get()]
        inputColEntryVar.set(columns[0])
        outputColEntryVar.set(columns[1])

    # style stuff
    fontType = "None"
    fontSize = 15
    fontGeneral = (fontType, fontSize)
    fontSmall = (fontType, int(fontSize/3*2))

    # root stuff
    root = tk.Tk()
    root.title(TITLE)
    rootWidth = 350
    rootHeight = 290
    root.geometry("{}x{}".format(rootWidth, rootHeight))
    root.resizable(0, 0)
    # root.attributes('-topmost', True)  # keeps root window at top layer

    # frame stuff        
    frames = []
    for i in range(7):
        frames.append(tk.Frame(root, bd=0, relief=tk.SOLID))
        frames[i].pack(padx=10, pady=3)  # fill="x", 

    # var stuff
    excelPathVar = tk.StringVar(value="~~~")
    inputColEntryVar = tk.StringVar()
    outputColEntryVar = tk.StringVar()
    firstRowVar = tk.IntVar(value=2)
    lastRowVar = tk.IntVar(value="")

    # data variables
    wb = None

    ## core UI elements stuff
    #
    browseButton = tk.Button(frames[0], text="Browse to Select", command=selectExcel, font=fontGeneral, width=rootWidth)
    excelPathLabel = tk.Label(frames[0], textvariable=excelPathVar, font=fontSmall, anchor="e")
    browseButton.pack()
    excelPathLabel.pack()

    #
    worksheetHeaderLabel = tk.Label(frames[1], text="Input Worksheet:", font=fontGeneral)
    worksheetCombobox = ttk.Combobox(frames[1], state="readonly")
    worksheetHeaderLabel.pack(side=tk.LEFT)
    worksheetCombobox.pack(side=tk.LEFT)

    #
    procedureHeaderLabel = tk.Label(frames[2], text="Procedure: ", font=fontGeneral)
    procedureCombobox = ttk.Combobox(frames[2], state="readonly", values=list(PROCEDURES_DISPLAY))
    procedureHeaderLabel.pack(side=tk.LEFT)
    procedureCombobox.pack(side=tk.LEFT)

    #
    inputColLabel = tk.Label(frames[3], text="Input column: ", font=fontGeneral)
    inputColEntry = tk.Entry(frames[3], textvariable=inputColEntryVar, font=fontSmall)
    inputColLabel.pack(side=tk.LEFT)
    inputColEntry.pack(side=tk.LEFT)

    #
    outputColLabel = tk.Label(frames[4], text="Output column: ", font=fontGeneral)
    outputColEntry = tk.Entry(frames[4], textvariable=outputColEntryVar, font=fontSmall)
    outputColLabel.pack(side=tk.LEFT)
    outputColEntry.pack(side=tk.LEFT)

    #
    rowLabel = tk.Label(frames[5], text="Row Range: ", font=fontGeneral)
    firstRowEntry = tk.Entry(frames[5], textvariable=firstRowVar, font=fontSmall, width=4)
    lastRowEntry = tk.Entry(frames[5], textvariable=lastRowVar, font=fontSmall, width=4)
    rowLabel.pack(side=tk.LEFT)
    firstRowEntry.pack(side=tk.LEFT)
    lastRowEntry.pack(side=tk.LEFT)

    #
    executeButton = tk.Button(frames[6], text="Execute", command=launchControllerFromView, font=fontGeneral)
    executeButton.pack()

    # Bindings
    worksheetCombobox.bind("<<ComboboxSelected>>", lambda _: updateMaxRow())
    procedureCombobox.bind("<<ComboboxSelected>>", lambda _: updateColumns())

    # color mode change
    guiC = guiController.GUIController(root)
    guiC.standardInitialize()


    root.mainloop()
