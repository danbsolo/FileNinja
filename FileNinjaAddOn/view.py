import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import ttk
from FileNinjaSuite.FileNinjaAddOn import control
import openpyxl as opxl
from FileNinjaSuite.FileNinjaAddOn.defs import *
from FileNinjaSuite.Shared import guiController


def launchView():
    def initiateControllerThread():
        guiC.setThreadVars(executeButton)
        guiC.initiateControllerThread(lambda:
                                        control.launchController(
                                        excelPathVar.get(),
                                        wb,
                                        worksheetCombobox.get(),
                                        procedureCombobox.get(),
                                        inputColEntryVar.get(),
                                        outputColEntryVar.get(),
                                        firstRowVar.get(),
                                        lastRowVar.get()))


    def selectExcel():
        excelPath = askopenfilename(title="Select an EXCEL workbook",
                            filetypes=[("Excel files", "*.xlsx *.xlsm")])
        
        if not excelPath: return
        
        excelPathVar.set(excelPath)

        nonlocal wb
        wb = opxl.load_workbook(excelPath)

        # Update the combobox accordingly
        worksheetCombobox.config(values=wb.sheetnames)
        worksheetCombobox.set(wb.sheetnames[0])
        worksheetCombobox.event_generate("<<ComboboxSelected>>") # or just call the function directly


    def updateMaxRow():
        lastRowVar.set(wb[worksheetCombobox.get()].max_row)

    def updateColumns():
        columns = PROCEDURES_DISPLAY[procedureCombobox.get()]
        inputColEntryVar.set(columns[0])
        outputColEntryVar.set(columns[1])

    # style stuff
    fontType = "None"
    fontSize = 15
    fontGeneral = (fontType, fontSize)
    fontSmall = (fontType, int(fontSize/3*2))

    # root stuff
    root = tk.Tk()
    root.title(TITLE)
    rootWidth = 350
    rootHeight = 290
    root.geometry("{}x{}".format(rootWidth, rootHeight))
    root.resizable(0, 0)
    # root.attributes('-topmost', True)  # keeps root window at top layer

    # frame stuff        
    frames = []
    for i in range(7):
        frames.append(tk.Frame(root, bd=0, relief=tk.SOLID))
        frames[i].pack(padx=10, pady=3)  # fill="x", 

    # var stuff
    excelPathVar = tk.StringVar(value="~~~")
    inputColEntryVar = tk.StringVar()
    outputColEntryVar = tk.StringVar()
    firstRowVar = tk.IntVar(value=2)
    lastRowVar = tk.IntVar(value="")

    # data variables
    wb = None

    ## core UI elements stuff
    #
    browseButton = tk.Button(frames[0], text="Browse to Select", command=selectExcel, font=fontGeneral, width=rootWidth)
    excelPathLabel = tk.Label(frames[0], textvariable=excelPathVar, font=fontSmall, anchor="e")
    browseButton.pack()
    excelPathLabel.pack()

    #
    worksheetHeaderLabel = tk.Label(frames[1], text="Input Worksheet:", font=fontGeneral)
    worksheetCombobox = ttk.Combobox(frames[1], state="readonly")
    worksheetHeaderLabel.pack(side=tk.LEFT)
    worksheetCombobox.pack(side=tk.LEFT)

    #
    procedureHeaderLabel = tk.Label(frames[2], text="Procedure: ", font=fontGeneral)
    procedureCombobox = ttk.Combobox(frames[2], state="readonly", values=list(PROCEDURES_DISPLAY))
    procedureHeaderLabel.pack(side=tk.LEFT)
    procedureCombobox.pack(side=tk.LEFT)

    #
    inputColLabel = tk.Label(frames[3], text="Input column: ", font=fontGeneral)
    inputColEntry = tk.Entry(frames[3], textvariable=inputColEntryVar, font=fontSmall)
    inputColLabel.pack(side=tk.LEFT)
    inputColEntry.pack(side=tk.LEFT)

    #
    outputColLabel = tk.Label(frames[4], text="Output column: ", font=fontGeneral)
    outputColEntry = tk.Entry(frames[4], textvariable=outputColEntryVar, font=fontSmall)
    outputColLabel.pack(side=tk.LEFT)
    outputColEntry.pack(side=tk.LEFT)

    #
    rowLabel = tk.Label(frames[5], text="Row Range: ", font=fontGeneral)
    firstRowEntry = tk.Entry(frames[5], textvariable=firstRowVar, font=fontSmall, width=4)
    lastRowEntry = tk.Entry(frames[5], textvariable=lastRowVar, font=fontSmall, width=4)
    rowLabel.pack(side=tk.LEFT)
    firstRowEntry.pack(side=tk.LEFT)
    lastRowEntry.pack(side=tk.LEFT)

    #
    executeButton = tk.Button(frames[6], text="Execute", command=initiateControllerThread, font=fontGeneral)
    executeButton.pack()

    # Bindings
    worksheetCombobox.bind("<<ComboboxSelected>>", lambda _: updateMaxRow())
    procedureCombobox.bind("<<ComboboxSelected>>", lambda _: updateColumns())

    # color mode change
    guiC = guiController.GUIController(root, TITLE)
    guiC.standardInitialize()


    root.mainloop()
